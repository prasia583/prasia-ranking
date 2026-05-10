ChatGPT


업로드한 이미지
무슨오류야?

업로드한 이미지

업로드한 이미지
ranking 2026 05 11레벨별통계


뭐가 문제인거지?

업로드한 이미지

붙여넣은 텍스트 (1)(13).txt
문서
이런구조고 니가 말한코드는


붙여넣은 텍스트 (1)(14).txt
문서
이렇게 하면 되는걸까?

아래처럼만 바꿔서 그대로 붙여넣으면 돼.

1. safe_int 함수 추가
safe_num 함수 바로 아래에 추가:

def safe_int(x, default=0):
    try:
        if x is None:
            return default

        if isinstance(x, (int, float)):
            return int(x)

        s = str(x).strip().replace(",", "")

        if s == "":
            return default

        return int(float(s))

    except Exception:
        return default
2. 아래 부분들 전부 교체
A. sort_member_rows
기존 코드:

def sort_member_rows(rows):
    return sorted(
        rows,
        key=lambda x: (
            -int(x.get("grade", 0) or 0),
            -int(x.get("level", 0) or 0),
            safe_str(x.get("guild", "")),
            safe_str(x.get("nickname", "")),
        )
    )
↓

교체:

def sort_member_rows(rows):
    return sorted(
        rows,
        key=lambda x: (
            -safe_int(x.get("grade", 0)),
            -safe_int(x.get("level", 0)),
            safe_str(x.get("guild", "")),
            safe_str(x.get("nickname", "")),
        )
    )
B. g["membersList"]
기존:

g["membersList"] = sorted(
    g["membersList"],
    key=lambda x: (
        -int(x.get("grade", 0) or 0),
        -int(x.get("level", 0) or 0),
        safe_str(x.get("nickname", "")),
    )
)
↓

교체:

g["membersList"] = sorted(
    g["membersList"],
    key=lambda x: (
        -safe_int(x.get("grade", 0)),
        -safe_int(x.get("level", 0)),
        safe_str(x.get("nickname", "")),
    )
)
C. g["byClassMembers"]
기존:

g["byClassMembers"] = {
    k: sorted(
        v,
        key=lambda x: (
            -int(x.get("grade", 0) or 0),
            -int(x.get("level", 0) or 0),
            safe_str(x.get("nickname", "")),
        )
    )
    for k, v in g["byClassMembers"].items()
}
↓

교체:

g["byClassMembers"] = {
    k: sorted(
        v,
        key=lambda x: (
            -safe_int(x.get("grade", 0)),
            -safe_int(x.get("level", 0)),
            safe_str(x.get("nickname", "")),
        )
    )
    for k, v in g["byClassMembers"].items()
}
D. g["byGradeMembers"]
기존:

g["byGradeMembers"] = {
    k: sorted(
        v,
        key=lambda x: (
            -int(x.get("grade", 0) or 0),
            -int(x.get("level", 0) or 0),
            safe_str(x.get("nickname", "")),
        )
    )
    for k, v in g["byGradeMembers"].items()
}
↓

교체:

g["byGradeMembers"] = {
    k: sorted(
        v,
        key=lambda x: (
            -safe_int(x.get("grade", 0)),
            -safe_int(x.get("level", 0)),
            safe_str(x.get("nickname", "")),
        )
    )
    for k, v in g["byGradeMembers"].items()
}
중요
위 3개(B,C,D)는 파일 안에 2번 존재해.

토벌상위분류 처리 부분 1세트

fallback 서버 시트 부분 1세트

즉 총:

A 1개

B 2개

C 2개

D 2개

수정해야 해.

가장 쉬운 방법
GitHub에서:

-int(
검색해서 전부:

-safe_int(
로 바꾸고,

혹시 이상하게 바뀐:

safe_int("grade"
safe_int("level"
이런 건 전부 삭제하면 돼.


한번에 붙여넣게 전체코드를 달라고 기존거 다 지우고

def safe_int(x, default=0):
    try:
        if x is None:
            return default

        if isinstance(x, (int, float)):
            return int(x)

        s = str(x).strip().replace(",", "")

        if s == "":
            return default

        return int(float(s))

    except Exception:
        return default
그리고 아래 코드를 기존 것들 전부 지우고 그대로 교체해.

1. sort_member_rows 전체 교체
def sort_member_rows(rows):
    return sorted(
        rows,
        key=lambda x: (
            -safe_int(x.get("grade", 0)),
            -safe_int(x.get("level", 0)),
            safe_str(x.get("guild", "")),
            safe_str(x.get("nickname", "")),
        )
    )
2. membersList 전체 교체 (2군데)
g["membersList"] = sorted(
    g["membersList"],
    key=lambda x: (
        -safe_int(x.get("grade", 0)),
        -safe_int(x.get("level", 0)),
        safe_str(x.get("nickname", "")),
    )
)
3. byClassMembers 전체 교체 (2군데)
g["byClassMembers"] = {
    k: sorted(
        v,
        key=lambda x: (
            -safe_int(x.get("grade", 0)),
            -safe_int(x.get("level", 0)),
            safe_str(x.get("nickname", "")),
        )
    )
    for k, v in g["byClassMembers"].items()
}
4. byGradeMembers 전체 교체 (2군데)
g["byGradeMembers"] = {
    k: sorted(
        v,
        key=lambda x: (
            -safe_int(x.get("grade", 0)),
            -safe_int(x.get("level", 0)),
            safe_str(x.get("nickname", "")),
        )
    )
    for k, v in g["byGradeMembers"].items()
}
수정 후:

Commit changes

Actions 자동 실행

초록 체크 확인

하면 돼.


아니 완전히 처음부터 끝까지 다 달라고

2초 동안 생각함
아래 코드로 tools/build_snapshots.py 안에 있는 기존 코드 전부 삭제하고 전체 붙여넣기 하면 돼.

import re
import json
import shutil
from datetime import datetime
from pathlib import Path
from collections import defaultdict

from openpyxl import load_workbook

UPLOAD_DIR = Path("uploads")
OUT_DIR = Path("site") / "snapshots"
OUT_DIR.mkdir(parents=True, exist_ok=True)

PREFERRED_SHEETS = ["통합정렬", "통합 정렬", "결사랭킹", "결사 랭킹"]

LEVEL_STAT_SHEETS = ["레벨별통계", "레벨별 통계"]
HUNT_STAT_SHEETS = ["토벌등급별통계", "토벌등급별 통계"]

MEMBER_DETAIL_SHEETS = ["토벌상위분류"]


def safe_str(x):
    if x is None:
        return ""
    return str(x).strip()


def safe_num(x):
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return x
    s = str(x).strip().replace(",", "")
    if s == "":
        return None
    try:
        if "." in s:
            return float(s)
        return int(s)
    except Exception:
        return s


def safe_int(x, default=0):
    try:
        if x is None:
            return default
        if isinstance(x, (int, float)):
            return int(x)
        s = str(x).strip().replace(",", "")
        if s == "":
            return default
        return int(float(s))
    except Exception:
        return default


def guess_label_from_filename(stem: str) -> str:
    parts = stem.replace("-", "_").split("_")
    nums = [p for p in parts if p.isdigit()]
    if len(nums) >= 3:
        y, m, d = nums[-3], nums[-2], nums[-1]
        if len(y) == 4:
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return stem


def guess_date_key_from_filename(stem: str) -> str:
    parts = stem.replace("-", "_").split("_")
    nums = [p for p in parts if p.isdigit()]
    if len(nums) >= 3:
        y, m, d = nums[-3], nums[-2], nums[-1]
        if len(y) == 4:
            return f"{y}_{m.zfill(2)}_{d.zfill(2)}"
    return stem


def pick_worksheet(wb):
    for name in PREFERRED_SHEETS:
        if name in wb.sheetnames:
            return wb[name], name
    return wb[wb.sheetnames[0]], wb.sheetnames[0]


def pick_sheet_by_candidates(wb, candidates):
    for name in candidates:
        if name in wb.sheetnames:
            return wb[name], name
    return None, None


def normalize_header(v):
    return safe_str(v).replace("\n", "").replace("\r", "").replace(" ", "")


def build_header_map(ws, header_row=1, max_cols=80):
    m = {}
    for col in range(1, max_cols + 1):
        v = ws.cell(row=header_row, column=col).value
        if v is None:
            continue
        key = normalize_header(v)

        if key in ("순위",):
            m["rank"] = col
        elif key in ("결사명", "결사"):
            m["guild"] = col
        elif key in ("서버명", "서버"):
            m["server"] = col
        elif key in ("총점", "총점수", "총점수합계", "총점수(합계)"):
            m["total_score"] = col
        elif key in (
            "토벌등급점수",
            "토벌등급합계",
            "토벌등급점수합계",
            "토벌등급점수(합계)",
            "토벌등급합계점수",
        ):
            m["hunt_score"] = col
        elif key in (
            "레벨별점수",
            "레벨별합계",
            "레벨별점수합계",
            "레벨별점수(합계)",
            "레벨별합계점수",
        ):
            m["level_score"] = col

    return m


def find_server_sheet_header_map(ws, header_row=1, max_cols=200):
    colmap = {}
    for c in range(1, max_cols + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is None:
            continue
        key = normalize_header(v)

        if key in ("닉네임", "캐릭터명", "이름", "명칭"):
            colmap["nickname"] = c
        elif key in ("결사명", "결사"):
            colmap["guild"] = c
        elif key in ("클래스", "직업"):
            colmap["clazz"] = c
        elif key in ("토벌등급", "등급"):
            colmap["grade"] = c
        elif key in ("레벨", "Lv", "LV"):
            colmap["level"] = c

    return colmap


def is_server_sheet(ws) -> bool:
    hm = find_server_sheet_header_map(ws)
    return "guild" in hm


def build_member_count_map(wb):
    member_map = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name in PREFERRED_SHEETS:
            continue
        if sheet_name in LEVEL_STAT_SHEETS:
            continue
        if sheet_name in HUNT_STAT_SHEETS:
            continue
        if sheet_name in MEMBER_DETAIL_SHEETS:
            continue

        if not is_server_sheet(ws):
            continue

        hm = find_server_sheet_header_map(ws)
        gcol = hm["guild"]

        for r in range(2, ws.max_row + 1):
            guild = safe_str(ws.cell(row=r, column=gcol).value)

            if guild == "":
                continue
            if guild.replace(" ", "") == "-":
                continue

            key = (guild, sheet_name)
            member_map[key] = member_map.get(key, 0) + 1

    return member_map


def parse_guild_ranking(ws, member_map=None):
    header = build_header_map(ws, header_row=1)

    rank_col = header.get("rank", 1)
    guild_col = header.get("guild", 2)
    server_col = header.get("server", 3)
    hunt_col = header.get("hunt_score", 4)
    level_col = header.get("level_score", 5)
    total_col = header.get("total_score", 6)

    rows = []
    for r in range(2, ws.max_row + 1):
        rank = ws.cell(row=r, column=rank_col).value
        guild = ws.cell(row=r, column=guild_col).value
        server = ws.cell(row=r, column=server_col).value

        guild_str = safe_str(guild)
        server_str = safe_str(server)

        if rank is None and guild_str == "" and server_str == "":
            continue

        if guild_str.replace(" ", "") == "-":
            continue

        hunt_score = ws.cell(row=r, column=hunt_col).value
        level_score = ws.cell(row=r, column=level_col).value
        total_score = ws.cell(row=r, column=total_col).value

        members = 0
        if member_map is not None:
            members = member_map.get((guild_str, server_str), 0)

        rows.append({
            "rank": safe_num(rank),
            "guild": guild_str,
            "server": server_str,
            "members": members,
            "hunt_score": safe_num(hunt_score),
            "level_score": safe_num(level_score),
            "total_score": safe_num(total_score),
        })

    def sort_key(x):
        v = x.get("rank")
        if isinstance(v, (int, float)):
            return v
        return safe_int(v, 10**9)

    rows.sort(key=sort_key)

    for i, row in enumerate(rows, start=1):
        row["rank"] = i

    return rows


def parse_stat_sheet(ws, key_name):
    rows = []

    for r in range(2, ws.max_row + 1):
        key_val = ws.cell(row=r, column=1).value
        count_val = ws.cell(row=r, column=2).value
        ratio_val = ws.cell(row=r, column=3).value

        if key_val is None and count_val is None and ratio_val is None:
            continue

        key_str = safe_str(key_val)
        count_num = safe_num(count_val)
        ratio_num = safe_num(ratio_val)

        if key_str == "" and count_num is None:
            continue

        rows.append({
            key_name: key_str,
            "count": count_num if count_num is not None else 0,
            "ratio": ratio_num if isinstance(ratio_num, (int, float)) else 0,
        })

    def sort_key(x):
        return -safe_int(x.get(key_name), -999999)

    rows.sort(key=sort_key)
    return rows


def sort_member_rows(rows, include_guild=False):
    if include_guild:
        return sorted(
            rows,
            key=lambda x: (
                -safe_int(x.get("grade", 0)),
                -safe_int(x.get("level", 0)),
                safe_str(x.get("guild", "")),
                safe_str(x.get("nickname", "")),
            )
        )

    return sorted(
        rows,
        key=lambda x: (
            -safe_int(x.get("grade", 0)),
            -safe_int(x.get("level", 0)),
            safe_str(x.get("nickname", "")),
        )
    )


def parse_member_detail_sheet(ws):
    level_members = defaultdict(list)
    grade_members = defaultdict(list)

    for r in range(2, ws.max_row + 1):
        nickname = safe_str(ws.cell(row=r, column=1).value)
        guild = safe_str(ws.cell(row=r, column=2).value)
        clazz = safe_str(ws.cell(row=r, column=3).value)
        grade = safe_num(ws.cell(row=r, column=4).value)
        level = safe_num(ws.cell(row=r, column=5).value)
        server = safe_str(ws.cell(row=r, column=6).value)

        if nickname == "" and guild == "" and server == "":
            continue

        if server == "통합정렬":
            continue

        row = {
            "nickname": nickname,
            "guild": guild,
            "class": clazz,
            "grade": grade if isinstance(grade, (int, float)) else 0,
            "level": level if isinstance(level, (int, float)) else 0,
            "server": server,
        }

        level_key = str(safe_int(level)) if safe_int(level) > 0 else ""
        grade_key = str(safe_int(grade)) if safe_int(grade) > 0 else ""

        if level_key:
            level_members[level_key].append(row)
        if grade_key:
            grade_members[grade_key].append(row)

    level_members = {k: sort_member_rows(v, include_guild=True) for k, v in level_members.items()}
    grade_members = {k: sort_member_rows(v, include_guild=True) for k, v in grade_members.items()}

    return {
        "levelMembers": level_members,
        "huntGradeMembers": grade_members,
    }


def build_server_detail_data(wb):
    ws_member, _ = pick_sheet_by_candidates(wb, MEMBER_DETAIL_SHEETS)

    if ws_member:
        server_details = {}

        for r in range(2, ws_member.max_row + 1):
            nickname = safe_str(ws_member.cell(row=r, column=1).value)
            guild = safe_str(ws_member.cell(row=r, column=2).value)
            clazz = safe_str(ws_member.cell(row=r, column=3).value)
            grade = safe_num(ws_member.cell(row=r, column=4).value)
            level = safe_num(ws_member.cell(row=r, column=5).value)
            server = safe_str(ws_member.cell(row=r, column=6).value)

            if nickname == "" and guild == "" and server == "":
                continue

            if server == "통합정렬":
                continue

            if guild == "":
                continue
            if guild.replace(" ", "") == "-":
                continue
            if server == "":
                continue

            if server not in server_details:
                server_details[server] = {
                    "server": server,
                    "guilds": {}
                }

            guilds = server_details[server]["guilds"]

            if guild not in guilds:
                guilds[guild] = {
                    "members": 0,
                    "byClass": defaultdict(int),
                    "byGrade": defaultdict(int),
                    "membersList": [],
                    "byClassMembers": defaultdict(list),
                    "byGradeMembers": defaultdict(list),
                }

            member_row = {
                "nickname": nickname,
                "guild": guild,
                "class": clazz,
                "grade": grade if isinstance(grade, (int, float)) else 0,
                "level": level if isinstance(level, (int, float)) else 0,
                "server": server,
            }

            guilds[guild]["members"] += 1
            guilds[guild]["membersList"].append(member_row)

            if clazz != "":
                guilds[guild]["byClass"][clazz] += 1
                guilds[guild]["byClassMembers"][clazz].append(member_row)

            grade_key = str(safe_int(grade)) if safe_int(grade) > 0 else ""
            if grade_key != "":
                guilds[guild]["byGrade"][grade_key] += 1
                guilds[guild]["byGradeMembers"][grade_key].append(member_row)

        for server_name, server_data in server_details.items():
            for guild_name, g in server_data["guilds"].items():
                g["byClass"] = dict(sorted(g["byClass"].items(), key=lambda x: (-x[1], x[0])))

                def grade_sort_key(item):
                    return -safe_int(item[0], -999999)

                g["byGrade"] = dict(sorted(g["byGrade"].items(), key=grade_sort_key))

                g["membersList"] = sort_member_rows(g["membersList"])

                g["byClassMembers"] = {
                    k: sort_member_rows(v)
                    for k, v in g["byClassMembers"].items()
                }

                g["byGradeMembers"] = {
                    k: sort_member_rows(v)
                    for k, v in g["byGradeMembers"].items()
                }

        return server_details

    server_details = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        if sheet_name in PREFERRED_SHEETS:
            continue
        if sheet_name in LEVEL_STAT_SHEETS:
            continue
        if sheet_name in HUNT_STAT_SHEETS:
            continue
        if sheet_name in MEMBER_DETAIL_SHEETS:
            continue

        if not is_server_sheet(ws):
            continue

        hm = find_server_sheet_header_map(ws)
        gcol = hm.get("guild")
        ncol = hm.get("nickname")
        ccol = hm.get("clazz")
        grcol = hm.get("grade")
        lcol = hm.get("level")

        guilds = {}

        for r in range(2, ws.max_row + 1):
            guild = safe_str(ws.cell(row=r, column=gcol).value) if gcol else ""
            nickname = safe_str(ws.cell(row=r, column=ncol).value) if ncol else ""
            clazz = safe_str(ws.cell(row=r, column=ccol).value) if ccol else ""
            grade = safe_num(ws.cell(row=r, column=grcol).value) if grcol else 0
            level = safe_num(ws.cell(row=r, column=lcol).value) if lcol else 0

            if guild == "":
                continue
            if guild.replace(" ", "") == "-":
                continue

            if guild not in guilds:
                guilds[guild] = {
                    "members": 0,
                    "byClass": defaultdict(int),
                    "byGrade": defaultdict(int),
                    "membersList": [],
                    "byClassMembers": defaultdict(list),
                    "byGradeMembers": defaultdict(list),
                }

            member_row = {
                "nickname": nickname,
                "guild": guild,
                "class": clazz,
                "grade": grade if isinstance(grade, (int, float)) else 0,
                "level": level if isinstance(level, (int, float)) else 0,
                "server": sheet_name,
            }

            guilds[guild]["members"] += 1
            guilds[guild]["membersList"].append(member_row)

            if clazz != "":
                guilds[guild]["byClass"][clazz] += 1
                guilds[guild]["byClassMembers"][clazz].append(member_row)

            grade_key = str(safe_int(grade)) if safe_int(grade) > 0 else ""
            if grade_key != "":
                guilds[guild]["byGrade"][grade_key] += 1
                guilds[guild]["byGradeMembers"][grade_key].append(member_row)

        for guild_name, g in guilds.items():
            g["byClass"] = dict(sorted(g["byClass"].items(), key=lambda x: (-x[1], x[0])))

            def grade_sort_key(item):
                return -safe_int(item[0], -999999)

            g["byGrade"] = dict(sorted(g["byGrade"].items(), key=grade_sort_key))

            g["membersList"] = sort_member_rows(g["membersList"])

            g["byClassMembers"] = {
                k: sort_member_rows(v)
                for k, v in g["byClassMembers"].items()
            }

            g["byGradeMembers"] = {
                k: sort_member_rows(v)
                for k, v in g["byGradeMembers"].items()
            }

        server_details[sheet_name] = {
            "server": sheet_name,
            "guilds": guilds,
        }

    return server_details


def build_snapshots_from_uploads():
    index = []

    xlsx_files = sorted(list(UPLOAD_DIR.glob("*.xlsx")) + list(UPLOAD_DIR.glob("*.xlsm")))

    for xlsx_path in xlsx_files:
        stem = xlsx_path.stem
        label = guess_label_from_filename(stem)
        date_key = guess_date_key_from_filename(stem)

        ranking_out_name = f"{stem}.json"
        ranking_out_path = OUT_DIR / ranking_out_name

        stats_out_name = f"stats_{stem}.json"
        stats_out_path = OUT_DIR / stats_out_name

        member_stats_out_name = f"stats_members_{stem}.json"
        member_stats_out_path = OUT_DIR / member_stats_out_name

        detail_dir = OUT_DIR / f"detail_{date_key}"

        wb = load_workbook(xlsx_path, data_only=True)

        member_map = build_member_count_map(wb)

        ws_rank, used_sheet = pick_worksheet(wb)
        ranking_data = parse_guild_ranking(ws_rank, member_map)

        with open(ranking_out_path, "w", encoding="utf-8") as f:
            json.dump(ranking_data, f, ensure_ascii=False, indent=2)

        ws_level, level_sheet_name = pick_sheet_by_candidates(wb, LEVEL_STAT_SHEETS)
        ws_hunt, hunt_sheet_name = pick_sheet_by_candidates(wb, HUNT_STAT_SHEETS)

        stats_data = {
            "label": label,
            "file": ranking_out_name,
            "levelSheet": level_sheet_name,
            "huntSheet": hunt_sheet_name,
            "levelStats": parse_stat_sheet(ws_level, "level") if ws_level else [],
            "huntGradeStats": parse_stat_sheet(ws_hunt, "grade") if ws_hunt else [],
        }

        with open(stats_out_path, "w", encoding="utf-8") as f:
            json.dump(stats_data, f, ensure_ascii=False, indent=2)

        ws_member, member_sheet_name = pick_sheet_by_candidates(wb, MEMBER_DETAIL_SHEETS)
        member_stats_data = {
            "label": label,
            "file": ranking_out_name,
            "memberSheet": member_sheet_name,
            "levelMembers": {},
            "huntGradeMembers": {},
        }

        if ws_member:
            parsed_member_stats = parse_member_detail_sheet(ws_member)
            member_stats_data["levelMembers"] = parsed_member_stats["levelMembers"]
            member_stats_data["huntGradeMembers"] = parsed_member_stats["huntGradeMembers"]

        with open(member_stats_out_path, "w", encoding="utf-8") as f:
            json.dump(member_stats_data, f, ensure_ascii=False, indent=2)

        if detail_dir.exists():
            shutil.rmtree(detail_dir)
        detail_dir.mkdir(parents=True, exist_ok=True)

        server_details = build_server_detail_data(wb)

        for server_name, detail_data in server_details.items():
            file_server_name = safe_str(server_name).replace("\u00A0", " ").strip()

            name_candidates = {
                file_server_name,
                file_server_name.replace(" ", ""),
                file_server_name.replace(" ", "_"),
                file_server_name.replace(" ", "-"),
            }

            m = re.match(r"^(.*?)[\s_-]*(\d+)$", file_server_name)
            if m:
                base_name = safe_str(m.group(1)).replace("\u00A0", " ").strip()
                num_raw = m.group(2)
                num_no_zero = str(int(num_raw))
                num_two = num_no_zero.zfill(2)

                extra_candidates = {
                    f"{base_name}{num_no_zero}",
                    f"{base_name} {num_no_zero}",
                    f"{base_name}_{num_no_zero}",
                    f"{base_name}-{num_no_zero}",
                    f"{base_name}{num_two}",
                    f"{base_name} {num_two}",
                    f"{base_name}_{num_two}",
                    f"{base_name}-{num_two}",
                }

                name_candidates.update(extra_candidates)

            for candidate in name_candidates:
                if not candidate:
                    continue

                out_path = detail_dir / f"{candidate}.json"
                with open(out_path, "w", encoding="utf-8") as f:
                    json.dump(detail_data, f, ensure_ascii=False, indent=2)

        index.append({
            "label": label,
            "file": ranking_out_name,
            "statsFile": stats_out_name,
            "statsMembersFile": member_stats_out_name,
            "rows": len(ranking_data),
            "sheet": used_sheet,
        })

    def sort_key(item):
        try:
            return datetime.strptime(item["label"], "%Y-%m-%d")
        except Exception:
            return datetime.min

    index.sort(key=sort_key, reverse=True)

    with open(OUT_DIR / "index.json", "w", encoding="utf-8") as f:
        json.dump(index, f, ensure_ascii=False, indent=2)

    print(f"Generated {len(index)} snapshot(s) into {OUT_DIR}")


if __name__ == "__main__":
    build_snapshots_from_uploads()
기존 코드 기준으로 잘못 바뀐 safe_int("grade", 0) 부분까지 전부 정리한 전체본이야.


